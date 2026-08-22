#!/usr/bin/env python3
"""Transforme les bases Auschwitz en données statiques pour GitHub Pages.

Le site reste entièrement gratuit : la base SQL et le classeur maître sont
convertis en fichiers JSON découpés par corpus. Le navigateur ne télécharge
que la catégorie consultée, y compris pour le registre de 8 502 fiches SS.
"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time
from html import escape
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_BASE_URL = "https://victeams.github.io/Le-panthon-des-heros-aushwist-41-45"
DEFAULT_SQL = Path("sources/base-complete-auschwitz.sql")
DEFAULT_WORKBOOK = Path("sources/inventaire-auschwitz-base-maitre.xlsx")
DEFAULT_OUTPUT = Path("data/base-documentaire")
MAX_RECORDS_PER_FILE = 350


@dataclass(frozen=True)
class Section:
    slug: str
    title: str
    description: str
    source: str
    table: str | None = None
    sheet: str | None = None
    header_row: int = 1


SQL_SECTIONS = (
    Section("camps", "Camps principaux", "Auschwitz I, Birkenau et Monowitz.", "Base SQL", table="camps"),
    Section("vie-quotidienne", "Vie quotidienne", "Conditions de détention, faim, travail forcé, appels et résistance.", "Base SQL", table="daily_life"),
    Section("lieux-historiques", "Lieux historiques", "Installations, fonctions, périodes et devenir des lieux.", "Base SQL", table="facilities"),
    Section("femmes-31000", "Femmes du convoi des 31 000", "Liste nominative de travail avec matricules et état de vérification.", "Base SQL", table="french_31000"),
    Section("hommes-45000", "Hommes du convoi des 45 000", "Liste nominative de travail avec matricules et état de vérification.", "Base SQL", table="french_45000"),
    Section("personnes", "Personnes documentées", "Notices nominatives complémentaires présentes dans la base historique.", "Base SQL", table="people"),
    Section("verifications-nominatives", "Vérifications nominatives", "Journal des identités et matricules rapprochés des sources.", "Base SQL", table="nominative_verification_log"),
    Section("sous-camps", "Sous-camps", "Sous-camps du complexe, périodes, travail et effectifs documentés.", "Base SQL", table="subcamps"),
    Section("chronologie", "Chronologie", "Événements datés de l’histoire du complexe d’Auschwitz.", "Base SQL", table="timeline"),
    Section("transports", "Transports", "Origines, dates, effectifs et statuts documentaires des convois.", "Base SQL", table="transports"),
    Section("statistiques-victimes", "Statistiques des victimes", "Ordres de grandeur conservant leur niveau de précision.", "Base SQL", table="victim_statistics"),
    Section("origines-juives", "Origines des déportés juifs", "Estimations par pays ou origine, avec sources et niveaux de confiance.", "Base SQL", table="jewish_origins"),
    Section("experiences-medicales", "Expériences médicales", "Lieux, responsables, groupes victimes et conséquences documentées.", "Base SQL", table="medical_experiments"),
    Section("responsables", "Responsables", "Fonctions, périodes et devenir après-guerre des responsables documentés.", "Base SQL", table="officials"),
    Section("glossaire", "Glossaire historique", "Définitions des termes utilisés dans la base documentaire.", "Base SQL", table="glossary"),
    Section("sources", "Sources de la base", "Institutions, titres, liens et appréciation de fiabilité.", "Base SQL", table="sources"),
)


WORKBOOK_SECTIONS = (
    Section("batiments", "Bâtiments et installations", "Inventaire détaillé des bâtiments des trois camps principaux.", "Classeur maître", sheet="Bâtiments principaux", header_row=3),
    Section("direction", "Direction du complexe", "Chaîne de commandement et fonctions datées.", "Classeur maître", sheet="Direction", header_row=3),
    Section("personnel-cle", "Personnel de direction et fonctions clés", "Sélection structurante, non exhaustive, des fonctions documentées.", "Classeur maître", sheet="Personnel clé", header_row=3),
    Section("personnel-feminin", "Personnel féminin au service de la SS", "Surveillantes, opératrices radio et infirmières documentées.", "Classeur maître", sheet="Personnel féminin", header_row=3),
    Section("personnel-ss", "Registre du personnel SS", "Index public des 8 502 fiches du registre IPN, chargé uniquement sur demande.", "Classeur maître · registre IPN", sheet="Personnel SS", header_row=3),
    Section("documents-judiciaires", "Documents judiciaires liés au personnel SS", "Décisions et pièces judiciaires reliées aux notices IPN.", "Classeur maître · registre IPN", sheet="Documents judiciaires SS", header_row=3),
    Section("tatoueurs-detenus", "Tatoueurs détenus", "Identités documentées et limites des archives, sans les confondre avec le personnel SS.", "Classeur maître", sheet="Tatoueurs détenus", header_row=9),
    Section("lieux-tatouage", "Lieux du tatouage", "Lieux et périodes documentés du tatouage des matricules.", "Classeur maître", sheet="Lieux du tatouage", header_row=5),
    Section("securite-installations", "Sécurité et installations", "Clôtures, miradors, sanitaires, crématoires et autres installations documentées.", "Classeur maître", sheet="Sécurité et installations", header_row=10),
    Section("transferts-recents", "Transferts et ajouts récents", "Corpus récemment identifiés et pistes de recherche documentées.", "Classeur maître", sheet="Transferts & ajouts récents", header_row=1),
    Section("personnes-verifiees", "Personnes nouvellement vérifiées", "Identités récemment rapprochées de sources institutionnelles.", "Classeur maître", sheet="Personnes nouvellement vérifiées", header_row=1),
)


LABELS = {
    "camp_id": "Identifiant",
    "topic_id": "Identifiant",
    "facility_id": "Identifiant",
    "list_id": "Identifiant",
    "person_id": "Identifiant",
    "subcamp_id": "Identifiant",
    "event_id": "Identifiant",
    "transport_id": "Identifiant",
    "group_id": "Identifiant",
    "origin_id": "Identifiant",
    "experiment_id": "Identifiant",
    "official_id": "Identifiant",
    "term_id": "Identifiant",
    "source_id": "Source",
    "name": "Nom",
    "other_names": "Autres noms",
    "location_historic": "Lieu historique",
    "location_current": "Lieu actuel",
    "start": "Début",
    "end": "Fin",
    "primary_functions": "Fonctions principales",
    "key_places": "Lieux principaux",
    "notes": "Notes",
    "confidence": "Niveau de confiance",
    "theme": "Thème",
    "summary": "Résumé",
    "human_impact": "Conséquences humaines",
    "type": "Type",
    "period": "Période",
    "function": "Fonction",
    "fate": "Devenir",
    "surname": "Nom",
    "person_label": "Prénom(s) et nom d’usage",
    "matricule": "Matricule",
    "matricule_status": "Statut du matricule",
    "review_status": "État de vérification",
    "convoy": "Convoi",
    "departure_date": "Date de départ",
    "arrival_date": "Date d’arrivée",
    "source_url": "Lien source",
    "last_name": "Nom",
    "first_name": "Prénom",
    "birth_date": "Date de naissance",
    "birth_place": "Lieu de naissance",
    "nationality": "Nationalité",
    "persecution_category": "Catégorie de persécution",
    "role_memory": "Rôle mémoriel",
    "prisoner_number": "Matricule",
    "arrival_context": "Contexte d’arrivée",
    "camp_or_sector": "Camp ou secteur",
    "death_date": "Date de décès",
    "death_place": "Lieu de décès",
    "current_place": "Lieu actuel",
    "labor": "Travail forcé",
    "employer": "Employeur",
    "population_reference": "Effectif documenté",
    "population_date": "Date de l’effectif",
    "sex": "Sexe",
    "date_start": "Date de début",
    "date_end": "Date de fin",
    "date_label": "Date affichée",
    "category": "Catégorie",
    "title": "Titre",
    "description": "Description",
    "place": "Lieu",
    "origin": "Origine",
    "transit": "Transit",
    "deportee_group": "Groupe déporté",
    "count_departed": "Nombre au départ",
    "count_registered": "Nombre enregistré",
    "count_murdered_on_arrival": "Nombre assassiné à l’arrivée",
    "count_survivors_known": "Survivants connus / libérés",
    "count_later_deaths": "Décès ultérieurs du convoi",
    "count_fate_unknown": "Destin inconnu",
    "coverage": "Couverture",
    "source_institution": "Institution source",
    "source_title": "Titre de la source",
    "group_name": "Groupe",
    "deported_min": "Déportés — minimum",
    "deported_max": "Déportés — maximum",
    "registered_estimate": "Enregistrés — estimation",
    "murdered_min": "Victimes — minimum",
    "murdered_max": "Victimes — maximum",
    "immediate_murder_estimate": "Assassinés à l’arrivée — estimation",
    "unit": "Unité",
    "precision": "Précision",
    "country_or_origin": "Pays ou origine",
    "deported_estimate": "Déportés — estimation",
    "note": "Note",
    "responsible": "Responsable(s)",
    "victim_groups": "Groupes victimes",
    "purpose_claimed": "But prétendu",
    "methods": "Méthodes",
    "consequences": "Conséquences",
    "role": "Fonction",
    "postwar_fate": "Devenir après-guerre",
    "term": "Terme",
    "definition": "Définition",
    "institution": "Institution",
    "url": "Lien",
    "language": "Langue",
    "accessed": "Consulté le",
    "reliability": "Fiabilité",
}


def normalized(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", " ", text.casefold()).strip()


def public_value(value: object) -> str | int | float | bool | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def record_title(values: dict[str, object], fallback: str) -> str:
    lookup = {normalized(key): value for key, value in values.items() if value not in (None, "")}
    for last_key, first_key in (
        ("nom", "prenom"),
        ("nom", "prenom s et nom d usage"),
        ("last name", "first name"),
        ("surname", "person label"),
    ):
        last, first = lookup.get(last_key), lookup.get(first_key)
        if last and first:
            return f"{last} {first}".strip()
    for key in (
        "nom complet graphie source",
        "batiment installation",
        "nom du sous camp",
        "nom",
        "theme",
        "titre",
        "title",
        "terme",
        "term",
        "groupe",
        "group name",
        "pays ou origine",
        "country or origin",
    ):
        if lookup.get(key):
            return str(lookup[key])
    for value in values.values():
        if value not in (None, "") and not re.fullmatch(r"[A-Z0-9-]+", str(value)):
            return str(value)
    return fallback


def make_record(raw: dict[str, object], fallback: str) -> dict[str, object]:
    values: dict[str, object] = {}
    record_id = fallback
    for index, (key, value) in enumerate(raw.items()):
        value = public_value(value)
        if value in (None, ""):
            continue
        label = LABELS.get(key, key.replace("_", " ").capitalize() if "_" in key else key)
        if index == 0:
            record_id = str(value)
        values[label] = value
    return {"id": record_id, "title": record_title(values, fallback), "values": values}


def table_exists(database: sqlite3.Connection, table: str) -> bool:
    return database.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)
    ).fetchone() is not None


def sql_records(path: Path) -> dict[str, list[dict[str, object]]]:
    database = sqlite3.connect(":memory:")
    database.row_factory = sqlite3.Row
    database.executescript(path.read_text(encoding="utf-8"))
    result: dict[str, list[dict[str, object]]] = {}
    for section in SQL_SECTIONS:
        assert section.table
        if not table_exists(database, section.table):
            continue
        if section.slug == "transports":
            rows = database.execute(
                'SELECT t.*, s.institution AS source_institution, '
                's.title AS source_title, s.url AS source_url '
                'FROM transports AS t LEFT JOIN sources AS s ON s.source_id = t.source_id'
            ).fetchall()
        else:
            rows = database.execute(f'SELECT * FROM "{section.table}"').fetchall()
        result[section.slug] = [
            make_record(dict(row), f"{section.slug}-{index:05d}")
            for index, row in enumerate(rows, start=1)
        ]
    database.close()
    return result


def unique_headers(cells: list[object]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for index, cell in enumerate(cells, start=1):
        base = str(cell.value or f"Colonne {index}").strip()
        seen[base] = seen.get(base, 0) + 1
        result.append(base if seen[base] == 1 else f"{base} ({seen[base]})")
    return result


def workbook_records(path: Path) -> dict[str, list[dict[str, object]]]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    result: dict[str, list[dict[str, object]]] = {}
    for section in WORKBOOK_SECTIONS:
        assert section.sheet
        if section.sheet not in workbook.sheetnames:
            continue
        sheet = workbook[section.sheet]
        header_cells = list(sheet[section.header_row])
        headers = unique_headers(header_cells)
        records: list[dict[str, object]] = []
        for row_number in range(section.header_row + 1, sheet.max_row + 1):
            cells = [sheet.cell(row_number, column) for column in range(1, len(headers) + 1)]
            first = public_value(cells[0].value)
            if first in (None, "") or not re.match(r"^[A-Z][A-Z0-9-]*-\d+", str(first)):
                continue
            raw: dict[str, object] = {}
            for header, cell in zip(headers, cells):
                value = public_value(cell.value)
                if cell.hyperlink and cell.hyperlink.target:
                    value = cell.hyperlink.target
                raw[header] = value
            records.append(make_record(raw, f"{section.slug}-{row_number:05d}"))
        result[section.slug] = records
    workbook.close()
    return result


def write_json(path: Path, payload: object, pretty: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if pretty:
        content = json.dumps(payload, ensure_ascii=False, indent=2)
    else:
        content = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    path.write_text(content + "\n", encoding="utf-8")


DATABASE_CSS = """
:root{color-scheme:dark;--bg:#0b0d10;--panel:#15191f;--panel2:#1c222a;--text:#f4f1e8;--muted:#b9c0c8;--gold:#d4ad62;--line:#303844}
*{box-sizing:border-box}body{margin:0;background:radial-gradient(circle at 20% 0,#202630 0,transparent 35%),var(--bg);color:var(--text);font-family:Arial,Helvetica,sans-serif;line-height:1.55}a{color:var(--gold)}.hero{padding:48px 22px 34px;text-align:center;border-bottom:1px solid var(--line)}.eyebrow{color:var(--gold);font-size:.82rem;font-weight:700;letter-spacing:.17em;text-transform:uppercase}.hero h1{max-width:920px;margin:8px auto;font:clamp(2rem,5vw,3.6rem)/1.1 Georgia,serif}.intro{max-width:850px;margin:18px auto;color:var(--muted)}.nav{display:flex;justify-content:center;flex-wrap:wrap;gap:9px;margin-top:24px}.nav a{padding:9px 14px;border:1px solid #5b5140;border-radius:999px;text-decoration:none;font-weight:700}.stats{display:flex;justify-content:center;flex-wrap:wrap;gap:10px;margin-top:25px}.stat{padding:9px 15px;background:#11151a;border:1px solid var(--line);border-radius:999px}.stat strong{color:var(--gold)}main{max-width:1180px;margin:auto;padding:30px 20px 64px}.notice{padding:18px 20px;border-left:4px solid var(--gold);background:#171b21;border-radius:8px;color:var(--muted)}.tools{display:grid;grid-template-columns:minmax(220px,1fr) minmax(260px,2fr);gap:12px;position:sticky;top:0;z-index:4;padding:16px 0;background:linear-gradient(var(--bg) 82%,transparent)}label span{display:block;margin:0 0 5px;color:var(--muted);font-size:.86rem}select,input{width:100%;padding:13px 14px;border:1px solid #414b58;border-radius:10px;background:#101419;color:var(--text);font-size:1rem}.summary{margin:8px 2px 18px;color:var(--muted)}.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(315px,1fr));gap:15px}.record{padding:20px;background:linear-gradient(145deg,var(--panel2),var(--panel));border:1px solid var(--line);border-radius:13px;box-shadow:0 10px 25px #0003}.record h2{margin:0 0 13px;font:1.28rem/1.25 Georgia,serif}.record dl{margin:0}.record dt{margin-top:9px;color:var(--gold);font-size:.78rem;font-weight:700;text-transform:uppercase;letter-spacing:.04em}.record dd{margin:2px 0;color:#e4e7eb;overflow-wrap:anywhere}.record details{margin-top:12px}.record summary{color:var(--gold);cursor:pointer}.pagination{display:flex;justify-content:center;align-items:center;gap:12px;margin-top:25px}.pagination button{padding:10px 15px;border:1px solid #5b5140;border-radius:999px;background:#171b21;color:var(--gold);font-weight:700}.pagination button:disabled{opacity:.35}.empty{padding:35px;text-align:center;color:var(--muted);border:1px dashed var(--line);border-radius:12px}footer{padding:23px;text-align:center;border-top:1px solid var(--line);color:var(--muted);font-size:.9rem}@media(max-width:680px){.tools{grid-template-columns:1fr;position:static}.grid{grid-template-columns:1fr}.hero{padding-top:36px}}
""".strip()


DATABASE_SCRIPT = r"""
const select=document.getElementById('category');const search=document.getElementById('search');const grid=document.getElementById('records');const summary=document.getElementById('summary');const previous=document.getElementById('previous');const next=document.getElementById('next');const pageLabel=document.getElementById('page');const PAGE_SIZE=30;let manifest;let records=[];let filtered=[];let page=1;const normalize=value=>String(value??'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').toLowerCase();
function fieldNode(label,value){const box=document.createDocumentFragment();const dt=document.createElement('dt');dt.textContent=label;const dd=document.createElement('dd');const text=String(value);if(/^https?:\/\//i.test(text)){const link=document.createElement('a');link.href=text;link.target='_blank';link.rel='noopener';link.textContent='Consulter la source';dd.append(link);}else{dd.textContent=text;}box.append(dt,dd);return box;}
function recordNode(record){const article=document.createElement('article');article.className='record';const title=document.createElement('h2');title.textContent=record.title;article.append(title);const entries=Object.entries(record.values);const visible=document.createElement('dl');entries.slice(0,6).forEach(([label,value])=>visible.append(fieldNode(label,value)));article.append(visible);if(entries.length>6){const details=document.createElement('details');const label=document.createElement('summary');label.textContent=`Afficher ${entries.length-6} information${entries.length-6>1?'s':''} supplémentaire${entries.length-6>1?'s':''}`;const rest=document.createElement('dl');entries.slice(6).forEach(([name,value])=>rest.append(fieldNode(name,value)));details.append(label,rest);article.append(details);}return article;}
function render(){grid.replaceChildren();const totalPages=Math.max(1,Math.ceil(filtered.length/PAGE_SIZE));page=Math.min(page,totalPages);const start=(page-1)*PAGE_SIZE;filtered.slice(start,start+PAGE_SIZE).forEach(record=>grid.append(recordNode(record)));if(!filtered.length){const empty=document.createElement('p');empty.className='empty';empty.textContent='Aucun résultat dans cette catégorie.';grid.append(empty);}summary.textContent=`${filtered.length.toLocaleString('fr-FR')} résultat${filtered.length>1?'s':''} — ${select.selectedOptions[0]?.textContent||''}`;pageLabel.textContent=`Page ${page} sur ${totalPages}`;previous.disabled=page<=1;next.disabled=page>=totalPages;}
function filter(){const query=normalize(search.value.trim());filtered=!query?records:records.filter(record=>normalize(`${record.title} ${Object.values(record.values).join(' ')}`).includes(query));page=1;render();}
async function loadCategory(){const category=manifest.categories.find(item=>item.slug===select.value);summary.textContent='Chargement…';grid.replaceChildren();const files=category.files||[category.file];const payloads=await Promise.all(files.map(async file=>{const response=await fetch(file);if(!response.ok)throw new Error(`Erreur ${response.status}`);return response.json();}));records=payloads.flat();filtered=records;search.value='';page=1;render();}
fetch('data/base-documentaire/manifest.json').then(response=>response.json()).then(data=>{manifest=data;document.getElementById('total').textContent=manifest.record_count.toLocaleString('fr-FR');document.getElementById('categories').textContent=manifest.categories.length.toLocaleString('fr-FR');manifest.categories.forEach(category=>{const option=document.createElement('option');option.value=category.slug;option.textContent=`${category.title} (${category.count.toLocaleString('fr-FR')})`;select.append(option);});return loadCategory();}).catch(error=>{summary.textContent=`Impossible de charger la base : ${error.message}`;});
select.addEventListener('change',()=>loadCategory().catch(error=>{summary.textContent=`Impossible de charger cette catégorie : ${error.message}`;}));search.addEventListener('input',filter);previous.addEventListener('click',()=>{page--;render();window.scrollTo({top:document.querySelector('.tools').offsetTop,behavior:'smooth'});});next.addEventListener('click',()=>{page++;render();window.scrollTo({top:document.querySelector('.tools').offsetTop,behavior:'smooth'});});
""".strip()


def database_page(base_url: str, record_count: int, category_count: int) -> str:
    canonical = f"{base_url.rstrip('/')}/base-documentaire.html"
    description = (
        "Base documentaire consultable sur Auschwitz : camps, bâtiments, convois, "
        "personnes, chronologie, sources et personnel documenté."
    )
    structured = json.dumps(
        {
            "@context": "https://schema.org",
            "@type": "Dataset",
            "name": "Base documentaire d’Auschwitz — Le Panthéon des héros",
            "description": description,
            "url": canonical,
            "inLanguage": "fr",
            "version": "2026-08-22",
            "isAccessibleForFree": True,
        },
        ensure_ascii=False,
        separators=(",", ":"),
    ).replace("</", "<\\/")
    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <meta name="description" content="{escape(description, quote=True)}">
  <meta name="robots" content="index,follow">
  <link rel="canonical" href="{escape(canonical, quote=True)}">
  <title>Base documentaire d’Auschwitz — Le Panthéon des héros</title>
  <script type="application/ld+json">{structured}</script>
  <style>{DATABASE_CSS}</style>
</head>
<body>
  <header class="hero">
    <p class="eyebrow">Archives • Sources • Mémoire</p>
    <h1>Base documentaire d’Auschwitz</h1>
    <p class="intro">Explorez les camps, bâtiments, convois, personnes, chronologies et sources réunis pour le projet. La recherche fonctionne directement dans votre navigateur, sans serveur payant.</p>
    <nav class="nav" aria-label="Navigation principale"><a href="index.html">Accueil</a><a href="femmes.html">Femmes 31000</a><a href="hommes.html">Hommes 45000</a><a href="photos.html">Photothèque</a></nav>
    <div class="stats"><span class="stat"><strong id="total">{record_count:,}</strong> notices</span><span class="stat"><strong id="categories">{category_count}</strong> catégories</span><span class="stat"><strong>0 €</strong> d’hébergement supplémentaire</span></div>
  </header>
  <main>
    <p class="notice"><strong>Précaution historique :</strong> les chiffres établis, minimums documentés, estimations et informations non attribuables restent distincts. Les lignes ne doivent pas être additionnées sans vérifier leur période, leur périmètre et leur source.</p>
    <div class="tools">
      <label><span>Catégorie documentaire</span><select id="category" aria-label="Choisir une catégorie"></select></label>
      <label><span>Rechercher dans la catégorie</span><input id="search" type="search" placeholder="Nom, matricule, lieu, date…" autocomplete="off"></label>
    </div>
    <p class="summary" id="summary" aria-live="polite">Chargement de la base…</p>
    <section class="grid" id="records" aria-label="Résultats documentaires"></section>
    <nav class="pagination" aria-label="Pagination"><button id="previous" type="button">← Précédent</button><span id="page">Page 1</span><button id="next" type="button">Suivant →</button></nav>
  </main>
  <footer>Base de travail historique — conserver les sources, les dates et les niveaux de certitude.</footer>
  <script>{DATABASE_SCRIPT}</script>
</body>
</html>
""".replace(f"{record_count:,}", f"{record_count:,}".replace(",", " "))


def build_database(root: Path, sql_path: Path, workbook_path: Path, output_dir: Path, base_url: str) -> dict[str, object]:
    root = root.resolve()
    sql_path = (root / sql_path).resolve() if not sql_path.is_absolute() else sql_path.resolve()
    workbook_path = (root / workbook_path).resolve() if not workbook_path.is_absolute() else workbook_path.resolve()
    output_dir = (root / output_dir).resolve() if not output_dir.is_absolute() else output_dir.resolve()
    if not sql_path.is_file():
        raise FileNotFoundError(f"Base SQL introuvable : {sql_path}")
    records_by_slug = sql_records(sql_path)
    if workbook_path.is_file():
        records_by_slug.update(workbook_records(workbook_path))
    else:
        manifest_path = output_dir / "manifest.json"
        if manifest_path.is_file():
            previous = json.loads(manifest_path.read_text(encoding="utf-8"))
            workbook_slugs = {section.slug for section in WORKBOOK_SECTIONS}
            for category in previous.get("categories", []):
                if category.get("slug") not in workbook_slugs:
                    continue
                files = category.get("files") or [category.get("file")]
                rows: list[dict[str, object]] = []
                for filename in files:
                    if not filename:
                        continue
                    path = root / str(filename)
                    if path.is_file():
                        rows.extend(json.loads(path.read_text(encoding="utf-8")))
                if rows:
                    records_by_slug[str(category["slug"])] = rows
            print("Classeur maître absent : conservation des catégories déjà générées.")

    definitions = {section.slug: section for section in (*SQL_SECTIONS, *WORKBOOK_SECTIONS)}
    output_dir.mkdir(parents=True, exist_ok=True)
    for old_json in output_dir.glob("*.json"):
        old_json.unlink()
    categories: list[dict[str, object]] = []
    for slug, records in records_by_slug.items():
        if not records:
            continue
        section = definitions[slug]
        category: dict[str, object] = {
            "slug": slug,
            "title": section.title,
            "description": section.description,
            "count": len(records),
            "source": section.source,
        }
        if len(records) <= MAX_RECORDS_PER_FILE:
            filename = f"{slug}.json"
            write_json(output_dir / filename, records)
            category["file"] = f"data/base-documentaire/{filename}"
        else:
            files: list[str] = []
            for index, start in enumerate(range(0, len(records), MAX_RECORDS_PER_FILE), start=1):
                filename = f"{slug}-{index:03d}.json"
                write_json(output_dir / filename, records[start : start + MAX_RECORDS_PER_FILE])
                files.append(f"data/base-documentaire/{filename}")
            category["files"] = files
        categories.append(category)

    record_count = sum(int(category["count"]) for category in categories)
    manifest: dict[str, object] = {
        "title": "Base documentaire d’Auschwitz",
        "version": "2026-08-22",
        "record_count": record_count,
        "category_count": len(categories),
        "methodology": "Les niveaux de certitude et les liens de source sont conservés. Les corpus qui se chevauchent ne doivent pas être additionnés.",
        "categories": categories,
    }
    write_json(output_dir / "manifest.json", manifest, pretty=True)
    (root / "base-documentaire.html").write_text(
        database_page(base_url, record_count, len(categories)), encoding="utf-8"
    )
    print(f"Base documentaire construite : {record_count} notices dans {len(categories)} catégories.")
    return manifest


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--sql", type=Path, default=DEFAULT_SQL)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL)
    arguments = parser.parse_args(argv)
    build_database(arguments.root, arguments.sql, arguments.workbook, arguments.output, arguments.base_url)
    return 0


if __name__ == "__main__":
    sys.exit(main())
